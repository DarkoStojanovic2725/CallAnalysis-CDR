import openpyxl
import datetime
import math
import xlsxwriter
# import tkinter as tk
# from tkinter import simpledialog
# root = tk.Tk()

class Interval:
    def __init__(self, beginTime, endTime, calls):
        self.beginTime = beginTime
        self.endTime = endTime
        self.calls = calls

    def __str__(self):
        return "%s, %s, %s" % (self.beginTime, self.endTime, self.calls)


class Call:
    def __init__(self, beginTime, duration):
        self.beginTime = beginTime
        self.duration = duration

    def __str__(self):
        return "%s, %s" % (self.beginTime, self.duration)


# opening CDR for reading
workbook = openpyxl.load_workbook("./Files/CDR_sample.xlsx")
sheet = workbook["Call_Records_from_2014-04-21_00"]
rowMax = sheet.max_row
columnMax = sheet.max_column

timeInterval = input("Enter time interval in minutes: ")
# timeInterval = simpledialog.askstring("Time interval", "Enter time interval in minutes:")
# creating output doc using xlsxwriter
workbookOutput = xlsxwriter.Workbook("./Files/Output.xlsx")
worksheet = workbookOutput.add_worksheet()
worksheet.set_column(0,2,20)
bold = workbookOutput.add_format({'bold': True})
worksheet.write('A1', 'Date', bold)
worksheet.write('B1', 'Start time', bold)
worksheet.write('C1', 'End time', bold)
worksheet.write('D1', '# of calls', bold)
chart = workbookOutput.add_chart({'type':'column'})
time_format = workbookOutput.add_format({'num_format': 'hh:mm:ss', 'align': 'center'})
time_format.set_align('center')
time_format.set_align('vcenter')
date_format = workbookOutput.add_format({'num_format': 'dd/mm/yyy', 'align': 'center'})
date_format.set_align('center')
date_format.set_align('vcenter')


numMinutesDelta = datetime.timedelta(minutes=int(timeInterval))

#filling in callsArray
callsArray = []

for row in range(2, rowMax + 1):
    call = Call(0, 0)
    for column in "EG":
        cell_name = "{}{}".format(column, row)
        if column == 'E':
            call.beginTime = sheet[cell_name].value
        if column == 'G':
            call.duration = sheet[cell_name].value
    callsArray.append(call)

intervalArray = []
foo = Interval(callsArray[0].beginTime, callsArray[0].beginTime + numMinutesDelta, 0)
intervalArray.append(foo)


startTime = callsArray[0].beginTime
endTime = callsArray[-1].beginTime + datetime.timedelta(seconds=int(callsArray[-1].duration))
passedTime = endTime-startTime
numberOfIntervals = math.ceil(passedTime.total_seconds() / 60 / int(timeInterval))+1
print("Number of intervals: ", numberOfIntervals)

# filling in intervalArray
for index in range(1, numberOfIntervals):
    intervalInstance = Interval(intervalArray[index - 1].endTime, intervalArray[index - 1].endTime + numMinutesDelta, 0)
    intervalArray.append(intervalInstance)

#to determine which interval the call falls into, checks ongoing calls from previous intervals
for i in range(len(callsArray)):
    endTime = callsArray[i].beginTime + datetime.timedelta(seconds=int(callsArray[i].duration))
    for y in range(len(intervalArray)):
        if intervalArray[y].beginTime <= callsArray[i].beginTime < intervalArray[y].endTime:
            intervalArray[y].calls += 1
            if endTime > intervalArray[y+1].beginTime and y != len(intervalArray)-1:
                intervalArray[y+1].calls += 1


for i in range(len(intervalArray)):
    print(intervalArray[i])


rowForWriting = 1
colForWriting = 0

for write in range(len(intervalArray)):
     worksheet.write_datetime(rowForWriting, colForWriting, intervalArray[write].beginTime, date_format)
     worksheet.write_datetime(rowForWriting, colForWriting + 1, intervalArray[write].beginTime, time_format)
     worksheet.write_datetime(rowForWriting, colForWriting + 2, intervalArray[write].endTime, time_format)
     worksheet.write_number(rowForWriting, colForWriting + 3, intervalArray[write].calls)
     rowForWriting += 1

values = '=' + worksheet.name+f'!$D2:$D{numberOfIntervals+1}'
names = '=' + worksheet.name+f'!$B2:$B{numberOfIntervals+1}'
chart.add_series({'values': values, 'categories': names})
chart.set_y_axis({'name': 'Number of calls'})
chart.set_x_axis({'name': 'Interval'})
worksheet.insert_chart('F5', chart)

workbookOutput.close()